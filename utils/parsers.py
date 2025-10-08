import os
import json
import pandas as pd
import sys
import urllib.parse
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment


sys.path.append(os.path.dirname(os.path.dirname(__file__)))

# === Load config struktur field ===
with open("config/struktur_fields.json", "r", encoding="utf-8") as f:
    STRUCT_FIELDS = json.load(f)

with open("config/templates.json", "r", encoding="utf-8") as f:
    TEMPLATES = json.load(f)

BASE_DIR = os.path.dirname(os.path.dirname(__file__))  # root project
CLEAN_DIR = os.path.join(BASE_DIR, "output", "cleaned")
OUT_DIR = os.path.join(BASE_DIR, "output", "parsed_output")
MSG_DIR = os.path.join(BASE_DIR, "output", "messages")
os.makedirs(OUT_DIR, exist_ok=True)
os.makedirs(MSG_DIR, exist_ok=True)


# ---- Helper ----
def latest_csv(prefix, path):
    files = [os.path.join(path, f) for f in os.listdir(path) if f.startswith(prefix) and f.endswith(".csv")]
    return max(files, key=os.path.getmtime) if files else None

def normalize_money(v):
    """Format angka jadi rupiah pakai titik sebagai pemisah ribuan"""
    if pd.isna(v) or v == "":
        return ""
    s = str(v).strip()
    clean = re.sub(r"[^\d]", "", s)  # hanya angka
    if not clean:
        return ""
    return f"{int(clean):,}".replace(",", ".")

def normalize_phone(v):
    """Pastikan nomor hp selalu diawali 0"""
    if pd.isna(v) or v == "":
        return ""
    s = str(v).strip()
    if s and not s.startswith("0"):
        s = "0" + s
    return s

def autosize_and_format_excel(path, table_name="DataTable"):
    """Atur lebar kolom, wrap text di kolom message, & jadikan tabel di Excel"""
    wb = load_workbook(path)
    ws = wb.active
    max_row = ws.max_row
    max_col = ws.max_column
    table_range = f"A1:{get_column_letter(max_col)}{max_row}"

    # Buat tabel resmi Excel
    tab = Table(displayName=table_name, ref=table_range)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Auto-resize kolom & wrap text khusus untuk kolom message
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        header = str(col[0].value).lower() if col[0].value else ""

        if header == "message":
            # wrap text untuk kolom message
            for cell in col:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.column_dimensions[col_letter].width = 60  # kasih lebar fix supaya enak dibaca
        else:
            # auto-width untuk kolom lain
            max_length = 0
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(path)

def normalize_column_names(df):
    """Normalisasi nama kolom dari lowercase ke UPPERCASE untuk template"""
    column_mapping = {}
    for col in df.columns:
        col_lower = col.lower()
        # Mapping dari lowercase ke UPPERCASE (untuk template)
        if col_lower == "no_sbg":
            column_mapping[col] = "NO_SBG"
        elif col_lower == "no_kredit":
            column_mapping[col] = "NO_KREDIT"
        elif col_lower == "nasabah":
            column_mapping[col] = "NASABAH"
        elif col_lower in ["telp_hp", "hp", "no_hp"]:
            column_mapping[col] = "TELP_HP"
        elif col_lower in ["tgl_jatuh_tempo", "tanggal_jatuh_tempo"]:
            column_mapping[col] = "TGL_JATUH_TEMPO"
        elif col_lower in ["tgl_kredit", "tanggal_kredit"]:
            column_mapping[col] = "TGL_KREDIT"
        elif col_lower == "uang_pinjaman":
            column_mapping[col] = "UANG_PINJAMAN"
        else:
            column_mapping[col] = col.upper()
    
    return df.rename(columns=column_mapping)

def generate_messages(df, doc_type):
    """Generate pesan WhatsApp sesuai template"""
    template = TEMPLATES.get(doc_type, "")
    messages = []

    # Normalisasi nama kolom ke UPPERCASE untuk template
    df = normalize_column_names(df)

    # SORTING: Urutkan berdasarkan UANG_PINJAMAN dari terbesar ke terkecil
    if "UANG_PINJAMAN" in df.columns:
        # Convert ke numeric untuk sorting yang benar
        df["UANG_PINJAMAN_NUMERIC"] = df["UANG_PINJAMAN"].apply(
            lambda x: int(re.sub(r"[^\d]", "", str(x))) if pd.notna(x) and str(x).strip() else 0
        )
        df = df.sort_values("UANG_PINJAMAN_NUMERIC", ascending=False)
        df = df.drop("UANG_PINJAMAN_NUMERIC", axis=1)
        print(f"[INFO] Messages diurutkan berdasarkan Uang Pinjaman (terbesar → terkecil)")

    for _, r in df.iterrows():
        row = r.to_dict()
        if "TELP_HP" in row:
            row["TELP_HP"] = normalize_phone(row["TELP_HP"])
        if "UANG_PINJAMAN" in row:
            row["UANG_PINJAMAN"] = normalize_money(row["UANG_PINJAMAN"])

        msg = template.format(**row)
        encoded = urllib.parse.quote(msg)
        hp = row.get("TELP_HP", "")
        wa_me = f"https://wa.me/{hp}?text={encoded}" if hp else ""
        wa_web = f"https://web.whatsapp.com/send?phone={hp}&text={encoded}" if hp else ""

        messages.append({"message": msg, "wa_me": wa_me, "wa_web": wa_web})

    return pd.DataFrame(messages)

def parse_document(doc_type):
    csv_file = latest_csv(doc_type, CLEAN_DIR)
    if not csv_file:
        print(f"[WARN] Tidak ada file CSV {doc_type} di {CLEAN_DIR}")
        return

    print(f"[INFO] Parsing {doc_type} dari {csv_file}")
    df = pd.read_csv(csv_file, encoding="utf-8-sig", dtype=str)

    # Normalisasi nama kolom ke UPPERCASE
    df = normalize_column_names(df)

    # Ambil field yang dibutuhkan sesuai config
    expected_fields = STRUCT_FIELDS.get(doc_type, [])
    
    # Filter hanya kolom yang ada di dataframe
    available_fields = [f for f in expected_fields if f in df.columns]
    if not available_fields:
        print(f"[ERROR] Tidak ada kolom yang sesuai dengan struktur {doc_type}")
        print(f"[INFO] Kolom yang tersedia: {list(df.columns)}")
        print(f"[INFO] Kolom yang diharapkan: {expected_fields}")
        return
    
    df_extracted = df[available_fields].copy()

    # normalisasi isi
    if "TELP_HP" in df_extracted.columns:
        df_extracted["TELP_HP"] = df_extracted["TELP_HP"].apply(normalize_phone)
    if "UANG_PINJAMAN" in df_extracted.columns:
        df_extracted["UANG_PINJAMAN"] = df_extracted["UANG_PINJAMAN"].apply(normalize_money)

    # simpan extracted ke CSV (parsed_output)
    out_csv = os.path.join(OUT_DIR, f"{doc_type}_extracted.csv")
    df_extracted.to_csv(out_csv, index=False, encoding="utf-8-sig")

    # simpan messages ke XLSX (SUDAH DISORT DI DALAM generate_messages)
    df_messages = generate_messages(df_extracted, doc_type)
    out_msg_xlsx = os.path.join(MSG_DIR, f"{doc_type}_messages.xlsx")
    df_messages.to_excel(out_msg_xlsx, index=False, engine="openpyxl")
    autosize_and_format_excel(out_msg_xlsx, f"{doc_type}_messages")

    print(f"[OK] Hasil {doc_type} tersimpan → extracted: {out_csv}, messages: {out_msg_xlsx}")
    return df_extracted.head()

def run_parsing(doc_type: str):
    return parse_document(doc_type)

if __name__ == "__main__":
    print(parse_document("jatuh_tempo"))
    print(parse_document("kredit_bermasalah"))